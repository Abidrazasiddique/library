from django.shortcuts import render
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm


# ================== Author Views ==================

class AuthorListView(ListView):
    model = Author
    template_name = "library/author_list.html"
    context_object_name = "authors"
    paginate_by = 10


class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = "library/author_form.html"
    success_url = reverse_lazy("library:author_list")

    def form_valid(self, form):
        messages.success(self.request, "Author added successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class AuthorUpdateView(UpdateView):
    model = Author
    form_class = AuthorForm
    template_name = "library/author_form.html"
    success_url = reverse_lazy("library:author_list")

    def form_valid(self, form):
        messages.success(self.request, "Author updated successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class AuthorDeleteView(DeleteView):
    model = Author
    template_name = "library/author_confirm_delete.html"
    success_url = reverse_lazy("library:author_list")

    def form_valid(self, form):
        messages.success(self.request, "Author deleted successfully!")
        return super().form_valid(form)


# ================== Book Views ==================

class BookListView(ListView):
    model = Book
    template_name = "library/book_list.html"
    context_object_name = "books"
    paginate_by = 10


class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = "library/book_form.html"
    success_url = reverse_lazy("library:book_list")

    def form_valid(self, form):
        messages.success(self.request, "Book added successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class BookUpdateView(UpdateView):
    model = Book
    form_class = BookForm
    template_name = "library/book_form.html"
    success_url = reverse_lazy("library:book_list")

    def form_valid(self, form):
        messages.success(self.request, "Book updated successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class BookDeleteView(DeleteView):
    model = Book
    template_name = "library/book_confirm_delete.html"
    success_url = reverse_lazy("library:book_list")

    def form_valid(self, form):
        messages.success(self.request, "Book deleted successfully!")
        return super().form_valid(form)


# ================== Borrow Record Views ==================

class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = "library/borrowrecord_list.html"
    context_object_name = "borrow_records"
    paginate_by = 10


class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = "library/borrowrecord_form.html"
    success_url = reverse_lazy("library:borrowrecord_list")

    def form_valid(self, form):
        messages.success(self.request, "Borrow record added successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class BorrowRecordUpdateView(UpdateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = "library/borrowrecord_form.html"
    success_url = reverse_lazy("library:borrowrecord_list")

    def form_valid(self, form):
        messages.success(self.request, "Borrow record updated successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class BorrowRecordDeleteView(DeleteView):
    model = BorrowRecord
    template_name = "library/borrowrecord_confirm_delete.html"
    success_url = reverse_lazy("library:borrowrecord_list")

    def form_valid(self, form):
        messages.success(self.request, "Borrow record deleted successfully!")
        return super().form_valid(form)


# ================== Export to Excel ==================

class ExportToExcelView(View):

    def get(self, request):
        workbook = Workbook()
        workbook.remove(workbook.active)

        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.create_author_sheet(workbook, header_font, center_align, border)
        self.create_book_sheet(workbook, header_font, center_align, border)
        self.create_borrow_sheet(workbook, header_font, center_align, border)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=library_data.xlsx"

        workbook.save(response)
        return response

    def create_author_sheet(self, wb, font, align, border):
        sheet = wb.create_sheet("Authors")
        headers = ["ID", "Name", "Email", "Bio"]

        for col, name in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=name)
            cell.font = font
            cell.alignment = align
            cell.border = border

        for row, author in enumerate(Author.objects.all(), 2):
            sheet.cell(row=row, column=1, value=author.id).border = border
            sheet.cell(row=row, column=2, value=author.name).border = border
            sheet.cell(row=row, column=3, value=author.email).border = border
            sheet.cell(row=row, column=4, value=author.bio).border = border

        self.adjust_width(sheet)

    def create_book_sheet(self, wb, font, align, border):
        sheet = wb.create_sheet("Books")
        headers = ["ID", "Title", "Genre", "Published Date", "Author"]

        for col, name in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=name)
            cell.font = font
            cell.alignment = align
            cell.border = border

        for row, book in enumerate(Book.objects.select_related("author"), 2):
            sheet.cell(row=row, column=1, value=book.id).border = border
            sheet.cell(row=row, column=2, value=book.title).border = border
            sheet.cell(row=row, column=3, value=book.get_genre_display()).border = border
            sheet.cell(row=row, column=4, value=str(book.published_date)).border = border
            sheet.cell(row=row, column=5, value=book.author.name).border = border

        self.adjust_width(sheet)

    def create_borrow_sheet(self, wb, font, align, border):
        sheet = wb.create_sheet("Borrow Records")
        headers = ["ID", "User", "Book", "Borrow Date", "Return Date", "Status"]

        for col, name in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=name)
            cell.font = font
            cell.alignment = align
            cell.border = border

        records = BorrowRecord.objects.select_related("book")
        for row, record in enumerate(records, 2):
            sheet.cell(row=row, column=1, value=record.id).border = border
            sheet.cell(row=row, column=2, value=record.user_name).border = border
            sheet.cell(row=row, column=3, value=record.book.title).border = border
            sheet.cell(row=row, column=4, value=str(record.borrow_date)).border = border
            sheet.cell(
                row=row,
                column=5,
                value=str(record.return_date) if record.return_date else "Not returned",
            ).border = border
            sheet.cell(
                row=row,
                column=6,
                value="Returned" if record.is_returned else "Borrowed",
            ).border = border

        self.adjust_width(sheet)

    def adjust_width(self, sheet):
        for column in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)


# ================== Home Page ==================

def home(request):
    context = {
        "author_count": Author.objects.count(),
        "book_count": Book.objects.count(),
        "borrow_count": BorrowRecord.objects.count(),
        "active_borrows": BorrowRecord.objects.filter(return_date__isnull=False).count(),
    }
    return render(request, "library/home.html", context)
